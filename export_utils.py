import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_excel(rows, filename_prefix):
    wb = Workbook()
    ws = wb.active
    ws.title = "Requests"
    ws.rtl = True

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A3D2F')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(horizontal='center', vertical='center')

    if rows:
        for col, header in enumerate(rows[0].keys(), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for r, row in enumerate(rows, 2):
            for c, (key, val) in enumerate(row.items(), 1):
                cell = ws.cell(row=r, column=c, value=str(val))
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            ws.column_dimensions[letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
